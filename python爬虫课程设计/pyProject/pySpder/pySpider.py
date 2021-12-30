import matplotlib
import openpyxl
import pymysql
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
import time
import requests
from lxml import etree
from multiprocessing import Pool,cpu_count,freeze_support
import matplotlib.pyplot as plt

#线程池是个什么东西，简单来说就是个分配任务的机器，不是说线程越多越好，有100个任务的话，那么开3个子线程，线程池是作用就是讲这100个任务丢进去让他运行，他们运行完之后就重新给他们分配任务，直到100个任务全部完成。
#开多个线程会消耗电脑的运存；
#引入线程池模块
from concurrent.futures import ThreadPoolExecutor
# from multiprocessing.dummy import Pool as ThreadPool

headers={
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36",
    "connection":"close"
}


datalist=[]  # 存储详情博客信息的列表 一篇博客作为一个元素存储到datalist中


#写入数据库操作
# 使用线程池写入，会出现pymysql.err.InternalError: Packet sequence number wrong - got 1 expected 2
# 原因：使用了多线程，多线程共享了同一个数据库连接，但每个execute前没有加上互斥锁
def writeIntoDataBase(datalist,spiderNum):
    db = pymysql.connect(host="localhost",user= "root", password= "wuwenqin", db="pythonspider") # 连接数据库
    cursor = db.cursor()
    i=0 # 进行定量数据存入
    for data in datalist:
        if i>=spiderNum:
            break
        sql="INSERT INTO spiderdata (title,author,content,keyword,publishedtime,linkurl,readcount) values ('%s','%s','%s','%s','%s','%s','%s')"%(pymysql.converters.escape_string(data[0]),pymysql.converters.escape_string(data[1]),pymysql.converters.escape_string(data[2]),pymysql.converters.escape_string(data[3]),pymysql.converters.escape_string(data[4]),pymysql.converters.escape_string(data[5]),pymysql.converters.escape_string(data[6]))
        cursor.execute(sql) #sql语句被执行
        i = i + 1
    db.commit()  # 写入数据，提交才能执行
    db.close()  # 查询完一定要关闭数据库啊
    # 多线程，记得关闭线程池
    # if nb_jobs >= 2:
    #     mpool.close()
    #     mpool.join()



#根据传入的详细博客地址进行数据爬取，每次存储一个博客文章的数据到datalist中
def Collect_information(url,key):
    only_1_page_data=[]
    # title="" # 标题
    content="" # 内容
    # author="" # 作者
    keyword=key # 关键字，根据关键字进行的查询
    # createtime="" # 创作时间
    response=requests.get(url=url,headers=headers).text
    rtree=etree.HTML(response)
    title=rtree.xpath('//*[@id="articleContentId"]/text()')[0] # 标题
    author= rtree.xpath('//*[@id="mainBox"]/main/div[1]/div[1]/div/div[2]/div[1]/div/a[1]/text()') # 作者
    print("文章标题："+title)
    createtime = rtree.xpath('//span[@class="time"]/text()')[0] # 创作时间
    contentList = rtree.xpath('//*[@id="article_content"]//text()') # 内容
    readcount=rtree.xpath('//*[@class="read-count"]/text()')[0] # 阅读量
    linkurl = url  # 链接地址
    contentList = [str(i) for i in contentList]
    # 转换成str字符串，否则还是list列表类型
    content=''.join(contentList)
    title=str(title)
    author=str(author)
    createtime=str(createtime)
    readcount=str(readcount)
    # 将only_1_page_data 存储到 datalist中
    only_1_page_data.append(title)
    only_1_page_data.append(author)
    only_1_page_data.append(content)
    only_1_page_data.append(keyword)
    only_1_page_data.append(createtime)
    only_1_page_data.append(linkurl)
    only_1_page_data.append(readcount)
    datalist.append(only_1_page_data)


# 根据传入的spiderNum爬取数量进行计算：需要滑动几次加载
def  cultByNum(web,spiderNum):
    initNum=30  # 一次加载大概30篇文章
    if spiderNum<initNum:
        pass
    else:
        needRoll=int(spiderNum/initNum)  # 向下取整
        if needRoll==1: # 只需滚动一次
            Rolldown(web)
        elif needRoll==2: # 两次
            Rolldown(web)
            Rolldown(web)
        else: # 根据needRoll的次数翻滚
            Rolldown(web)
            Rolldown(web)
            needRoll-=2
            while needRoll>0:
                needRoll-=1
                RollClick(web)

# 滚动到当前页面底部，一次加载30篇文章
def Rolldown(web):
    js = "window.scrollTo(0,document.body.scrollHeight)"
    web.execute_script(js)
    time.sleep(2)

# 滚动到底部，并点击加载更多按钮
def RollClick(web):
    Rolldown(web)  # 需要先滑动到底部，让加载按钮显示出来
    # js_top = "var q=document.documentElement.scrollTop=0"
    # web.execute_script(js_top)
    # js = "window.scrollTo(0,document.body.scrollHeight)"
    # web.execute_script(js)
    # time.sleep(2)
    web_Loadmore=web.find_element(By.XPATH,'//*[@id="app"]/div[2]/div[2]/div[1]/div[3]').click()
    time.sleep(2) # 休眠3秒，让数据能同步爬取


#在使用xlwt模块时，表的行数达到上限，该模块就无法使用了，会出现如题报错。故使用xlsx进行存储
def savedate(datalist,savepath,keyword,spiderNum):
    filename = '查询博客文章综合表-'+keyword+'.xlsx'
    savepath=savepath+"/"+filename  # 特定位置存储
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    # 表头
    header = ['标题', '作者', '内容', '关键字','上传时间','对应链接','阅读量']
    datalist.insert(0, header)
    i=0  # 进行数量判定
    for row in datalist:
        if i>=spiderNum:
            break
        outws.append(row)
        i=i+1
    outwb.save(savepath)  # 写入位置


# 这个对于大数据来说存储不了，有bug
# def savedate(datalist,savepath):
#     book=xlwt.Workbook(encoding="utf-8",style_compression=0)
#     sheet=book.add_sheet('查询博客文章综合表.xls',cell_overwrite_ok=True)
#     col=["标题","作者","内容","关键字","上传时间","对应链接"]#列表名称，根据需要对应几个写几个，对应下面的6
#     for i in range(0,6):#代表col的个数
#         sheet.write(0,i,col[i])#当i=0时，在第一行第一列写入"列1"
#     for data in datalist:#即将要写进去的最大行数（21，就是爬取的对象的有多少写多少）
#         i+=1
#         # date=datalist[i]
#         for j in range(0,6):#要写进去的列数
#             sheet.write(i+1,j,data[j])#当b=0，j=0中，在第二行，第1列中，写入要输入列表的第一个字符
#                                     # 当b=0，j=1中，在第二行，第1列中，写入要输入列表的第一个字符，直到第二行结束
#                                     # 当b=1，j=0中，在第二行，第1列中，写入要输入列表的第一个字符
#     book.save(savepath)


# 备份，这是使用进程池的任务执行方法
# # keyWord关键字搜索， spiderNum爬取数量,nb_jobs开启的线程数量
# def Web(keyWord,spiderNum,savepath,nb_jobs)  :
#     # 创建web-打开-登录-切换-输入-爬取
#     web = Chrome()  # 创建谷歌浏览器窗口；此处不适合用无头浏览器，因为这个网页最坑的就是登陆后要是没退出，一直会提醒那在其他地方登陆，这个又要重新打开，所以每次爬取完信息，都要自己手动关闭一下这个页面，否则下一次必定报错，需要手动把登陆信息给删除掉；
#     baseUrl="https://so.csdn.net/so/search?q="+keyWord +"&t=blog&u=&tm=365&urw=" # 拼接要查询的关键字
#     web.get(baseUrl)
#     time.sleep(3)
#     # Rolldown(web) # 滚动功能
#     # Rolldown(web) # 滚动功能
#     # RollClick(web)
#     cultByNum(web,spiderNum) # 根据传入的爬取数量进行滚动
#
#     # 通过xpath获取 与关键词相关的博客文章
#     # list=web.find_elements(By.XPATH,'//*[@id="app"]/div[2]/div[2]/div[1]/div[2]/div/div')
#     list=web.find_elements_by_class_name("list-item")
#     len=0  # 目前有多少博客文章
#
# # 多线程，判定是否需要多线程
#     if nb_jobs<2:
#         pass
#     else:
#         mpool=Pool(nb_jobs if nb_jobs else cpu_count())   # 进程池
#     # 通过链接，获取博客内的相关信息
#     linkedurlList=[]
#     for data in list:
#         href=data.find_element_by_class_name("block-title").get_attribute("href") # 获取链接
#         #通过链接获取信息，需要传入链接href以及关键字keyWord，进行后面的存储，在这里可以进行一个多线程
#         linkedurlList.append(href)
#         len = len + 1
#         print(href)
#
#     for url in linkedurlList:
#         if nb_jobs<2:
#             Collect_information(url,keyWord)
#         else: # 多线程提交任务
#             mpool.apply_async(Collect_information(url,keyWord))
#         # print(len)
#     # 多线程，记得关闭线程池
#     if nb_jobs >= 2:
#         mpool.close()
#         mpool.join()
#
#     writeIntoDataBase(datalist,spiderNum,nb_jobs) # 写入数据库中
#     savedate(datalist,savepath,keyWord,spiderNum)   #  存储到特定位置，以xlsx文件存储
#
#
#     return spiderNum


# keyWord关键字搜索， spiderNum爬取数量,nb_jobs开启的线程数量
def Web(keyWord,spiderNum,savepath,nb_jobs)  :
    datalist.clear() # 每次运行，清空原有数据
    print(keyWord,spiderNum,savepath,nb_jobs)
    # 创建web-打开-登录-切换-输入-爬取
    web = Chrome()  # 创建谷歌浏览器窗口；此处不适合用无头浏览器，因为这个网页最坑的就是登陆后要是没退出，一直会提醒那在其他地方登陆，这个又要重新打开，所以每次爬取完信息，都要自己手动关闭一下这个页面，否则下一次必定报错，需要手动把登陆信息给删除掉；
    baseUrl="https://so.csdn.net/so/search?q="+keyWord +"&t=blog&u=&tm=365&urw=" # 拼接要查询的关键字
    web.get(baseUrl)
    time.sleep(3)
    # Rolldown(web) # 滚动功能
    # Rolldown(web) # 滚动功能
    # RollClick(web)
    cultByNum(web,spiderNum) # 根据传入的爬取数量进行滚动

    # 通过xpath获取 与关键词相关的博客文章
    # list=web.find_elements(By.XPATH,'//*[@id="app"]/div[2]/div[2]/div[1]/div[2]/div/div')
    list=web.find_elements_by_class_name("list-item")
    curlen=0  # 目前有多少博客文章
    # 通过链接，获取博客内的相关信息
    linkedurlList=[]
    for data in list:
        href=data.find_element_by_class_name("block-title").get_attribute("href") # 获取链接
        #通过链接获取信息，需要传入链接href以及关键字keyWord，进行后面的存储，在这里可以进行一个多线程
        linkedurlList.append(href)
        curlen = curlen + 1
        # print(href)
    web.close()  # 关闭浏览器
# 线程池
    pool = ThreadPoolExecutor(max_workers=nb_jobs)
    for url in linkedurlList:
        if nb_jobs<2:
            print(url)
            Collect_information(url,keyWord)
        else: # 多线程提交任务
            pool.submit(Collect_information,url,keyWord)
    pool.shutdown(wait=True)
        # print(len)

    writeIntoDataBase(datalist,spiderNum) # 写入数据库中
    savedate(datalist,savepath,keyWord,spiderNum)   #  存储到特定位置，以xlsx文件存储
    return len(datalist)

# 根据关键字去数据库中查找，进行柱状图分析
def keyAnalysis():
    matplotlib.rcParams['font.sans-serif'] = ['KaiTi']
    x = []
    db = pymysql.connect(host="localhost", user="root", password="wuwenqin", db="pythonspider")  # 连接数据库
    cursor = db.cursor()
    select_sql="select DISTINCT keyword ,SUM(readcount) from spiderdata GROUP BY keyword"
    cursor.execute(select_sql)
    result=cursor.fetchall() #结果
    print(result)
    labels=[]
    x=[]
    for r in result:
        labels.append(r[0])
        x.append(r[1])
    print(labels)
    fig = plt.figure()
    plt.pie(x, labels=labels, autopct='%1.2f%%')  # 画饼图（数据，数据对应的标签，百分数保留两位小数点）
    plt.title("关键字阅读量占比图")
    plt.show()
    plt.savefig("keyAnalysis.jpg")

if __name__ == '__main__':
    keyAnalysis()
# if __name__ == '__main__':
#     keyWord="数据结构"
#     Web(keyWord,20)
#     time.sleep(5)                     #休眠2秒
    # url="https://blog.csdn.net/Strive_0902/article/details/84388684?ops_request_misc=%257B%2522request%255Fid%2522%253A%2522164015441316780265488735%2522%252C%2522scm%2522%253A%252220140713.130102334.pc%255Fall.%2522%257D&request_id=164015441316780265488735&biz_id=0&utm_medium=distribute.pc_search_result.none-task-blog-2~all~first_rank_ecpm_v1~rank_v31_ecpm-6-84388684.pc_search_insert_es_download_v2&utm_term=selenium%E7%9A%84%E6%BB%91%E5%8A%A8%E6%A8%A1%E5%9D%97%E8%A7%A3%E5%86%B3%E6%87%92%E5%8A%A0%E8%BD%BD&spm=1018.2226.3001.4187"
    # getBrief(url)
    # Collect_information(url,keyWord)






